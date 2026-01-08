from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.participant import Participant
from app.models.tour import Tour


def generate_tour_excel(tour: Tour, participants: List[Participant]) -> BytesIO:
    """
    Generate an Excel file for a tour with participant list and attendance checkbox.

    Args:
        tour: Tour object
        participants: List of participants for the tour

    Returns:
        BytesIO: Excel file in memory
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tour Participants"

    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Tour information section
    ws.merge_cells("A1:G1")
    tour_title = ws["A1"]
    tour_title.value = f"Tour: {tour.name}"
    tour_title.font = Font(bold=True, size=14)
    tour_title.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    tour_dates = ws["A2"]
    tour_dates.value = f"Dates: {tour.start_date} to {tour.end_date}"
    tour_dates.font = Font(size=11)
    tour_dates.alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:G3")
    tour_guides = ws["A3"]
    tour_guides.value = f"Guides: {tour.guides}"
    tour_guides.font = Font(size=11)
    tour_guides.alignment = Alignment(horizontal="center")

    # Add empty row
    ws.append([])

    # Headers
    headers = [
        "№",
        "Name",
        "Surname",
        "Birth Date",
        "Phone",
        "✓ Participated",
        "Signature",
    ]
    header_row = 5
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Participant data
    for idx, participant in enumerate(participants, 1):
        row_num = header_row + idx
        row_data = [
            idx,
            participant.name,
            participant.surname,
            str(participant.birth_date),
            participant.phone_number,
            "",  # Empty checkbox column
            "",  # Empty signature column
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Adjust column widths
    column_widths = {
        "A": 5,  # №
        "B": 15,  # Name
        "C": 15,  # Surname
        "D": 12,  # Birth Date
        "E": 15,  # Phone
        "F": 15,  # Participated
        "G": 20,  # Signature
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Set row heights
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 20

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return excel_file
